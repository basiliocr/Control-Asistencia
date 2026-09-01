from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .models import Pasante, Asistencia, Institucion, DiaEspecial, Correccion
from .forms import (
    PasanteForm,
    AsistenciaForm,
    DiaEspecialForm,
    RangoDiaEspecialForm,
    AdminUserForm,
)
from .services import registrar_con_gps, _calcular_tardanza


@login_required
def marcar(request):
    # Los administradores no marcan asistencia: van directo a su panel.
    if request.user.is_staff:
        return redirect("panel")
    if not hasattr(request.user, "pasante"):
        return render(request, "registro/marcar.html", {"sin_pasante": True})

    pasante = request.user.pasante
    resultado = None
    if request.method == "POST":
        tipo = request.POST.get("tipo")
        lat = request.POST.get("lat")
        lng = request.POST.get("lng")
        dispositivo = request.POST.get("dispositivo")
        resultado = registrar_con_gps(pasante, tipo, lat, lng, dispositivo)

    inst = Institucion.obtener()
    return render(
        request,
        "registro/marcar.html",
        {
            "pasante": pasante,
            "resultado": resultado,
            "institucion_lat": inst.latitud,
            "institucion_lng": inst.longitud,
            "radio": inst.radio_metros,
        },
    )


@staff_member_required
def panel(request):
    hoy = timezone.localdate()
    return render(
        request,
        "registro/panel.html",
        {
            "institucion": Institucion.obtener(),
            "total_pasantes": Pasante.objects.filter(activo=True).count(),
            "asistencias_hoy": Asistencia.objects.filter(fecha=hoy).count(),
            "tardanzas_hoy": Asistencia.objects.filter(
                fecha=hoy, tardanza_min__gt=0
            ).count(),
        },
    )


@staff_member_required
def ubicacion(request):
    inst = Institucion.obtener()
    if request.method == "POST":
        try:
            nombre = (request.POST.get("nombre") or "").strip()
            inst.nombre = nombre or inst.nombre
            inst.latitud = float(request.POST.get("latitud"))
            inst.longitud = float(request.POST.get("longitud"))
            inst.radio_metros = int(request.POST.get("radio_metros"))
            inst.save()
            messages.success(request, "Ubicación y perímetro actualizados.")
            return redirect("ubicacion")
        except (TypeError, ValueError):
            messages.error(request, "Los valores ingresados no son válidos.")
    return render(request, "registro/ubicacion.html", {"inst": inst})


# --------------------------- Pasantes ---------------------------


@staff_member_required
def pasantes_lista(request):
    pasantes = Pasante.objects.select_related("user").order_by("nombre")
    return render(request, "registro/pasantes_lista.html", {"pasantes": pasantes})


@staff_member_required
def pasante_form(request, pk=None):
    pasante = get_object_or_404(Pasante, pk=pk) if pk else None
    if request.method == "POST":
        form = PasanteForm(request.POST, instance=pasante)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            if pasante is None:
                if not password:
                    form.add_error(
                        "password", "La contraseña es obligatoria al crear un pasante."
                    )
                else:
                    user = User.objects.create_user(
                        username=username, password=password
                    )
                    user.is_active = form.cleaned_data["activo"]
                    user.save()
                    obj = form.save(commit=False)
                    obj.user = user
                    obj.save()
                    messages.success(request, f"Pasante «{obj.nombre}» creado.")
                    return redirect("pasantes_lista")
            else:
                user = pasante.user
                user.username = username
                if password:
                    user.set_password(password)
                user.is_active = form.cleaned_data["activo"]
                user.save()
                form.save()
                messages.success(request, f"Pasante «{pasante.nombre}» actualizado.")
                return redirect("pasantes_lista")
    else:
        form = PasanteForm(instance=pasante)
    return render(
        request, "registro/pasante_form.html", {"form": form, "pasante": pasante}
    )


@staff_member_required
def pasante_eliminar(request, pk):
    """Da de baja o reactiva un pasante. Ya NO elimina: el historial se conserva."""
    pasante = get_object_or_404(Pasante, pk=pk)
    if request.method == "POST":
        nuevo_estado = not pasante.activo  # alterna: baja <-> alta
        pasante.activo = nuevo_estado
        pasante.save(update_fields=["activo"])
        pasante.user.is_active = nuevo_estado
        pasante.user.save(update_fields=["is_active"])
        if nuevo_estado:
            messages.success(request, f"Pasante «{pasante.nombre}» reactivado.")
        else:
            messages.info(
                request,
                f"Pasante «{pasante.nombre}» dado de baja. Ya no podrá iniciar sesión; "
                f"su historial se conserva.",
            )
        return redirect("pasantes_lista")
    return render(request, "registro/pasante_eliminar.html", {"pasante": pasante})


# --------------------------- Planilla de asistencias ---------------------------


@staff_member_required
def asistencias_lista(request):
    # La lista solo se muestra DESPUÉS de filtrar (evita cargar todo de golpe).
    filtrado = bool(
        request.GET.get("desde")
        or request.GET.get("hasta")
        or request.GET.get("pasante") is not None
    )

    hoy = timezone.localdate()
    desde = request.GET.get("desde") or hoy.replace(day=1).isoformat()
    hasta = request.GET.get("hasta") or hoy.isoformat()
    pasante_id = request.GET.get("pasante") or ""

    asistencias = None
    if filtrado:
        asistencias = (
            Asistencia.objects.select_related("pasante")
            .filter(fecha__gte=desde, fecha__lte=hasta)
            .order_by("-fecha", "pasante__nombre")
        )
        if pasante_id:
            asistencias = asistencias.filter(pasante_id=pasante_id)

    return render(
        request,
        "registro/asistencias_lista.html",
        {
            "asistencias": asistencias,
            "pasantes": Pasante.objects.order_by("nombre"),
            "desde": desde,
            "hasta": hasta,
            "pasante_id": pasante_id,
            "filtrado": filtrado,
        },
    )


@staff_member_required
def asistencia_editar(request, pk):
    asistencia = get_object_or_404(Asistencia, pk=pk)
    if request.method == "POST":
        form = AsistenciaForm(request.POST, instance=asistencia)
        if form.is_valid():
            obj = form.save(commit=False)
            # La tardanza se recalcula sola según la hora de entrada
            # (aplicando tolerancia, feriados y horarios especiales).
            if obj.hora_entrada:
                obj.tardanza_min = _calcular_tardanza(
                    obj.pasante, obj.fecha, obj.hora_entrada
                )
            else:
                obj.tardanza_min = 0
            obj.save()
            if form.changed_data:
                Correccion.objects.create(
                    asistencia=asistencia,
                    admin=request.user,
                    detalle=(
                        f"Editado desde el panel. Campos: {', '.join(form.changed_data)}. "
                        f"Tardanza recalculada: {obj.tardanza_min} min."
                    ),
                )
            messages.success(
                request, "Asistencia actualizada. Se registró la corrección."
            )
            return redirect("asistencias_lista")
    else:
        form = AsistenciaForm(instance=asistencia)
    return render(
        request,
        "registro/asistencia_form.html",
        {"form": form, "asistencia": asistencia},
    )


@staff_member_required
def asistencia_eliminar(request, pk):
    asistencia = get_object_or_404(Asistencia, pk=pk)
    if request.method == "POST":
        asistencia.delete()
        messages.success(request, "Registro de asistencia eliminado.")
        return redirect("asistencias_lista")
    return render(
        request, "registro/asistencia_eliminar.html", {"asistencia": asistencia}
    )


# --------------------------- Días especiales ---------------------------


def _recalcular_tardanzas(fecha):
    """Recalcula la tardanza de todas las asistencias de una fecha.
    Se usa cuando se agrega, edita o quita un día especial que las afecta."""
    for a in Asistencia.objects.filter(fecha=fecha).select_related("pasante"):
        nueva = (
            _calcular_tardanza(a.pasante, a.fecha, a.hora_entrada)
            if a.hora_entrada
            else 0
        )
        if nueva != a.tardanza_min:
            a.tardanza_min = nueva
            a.save(update_fields=["tardanza_min"])


@staff_member_required
def dias_lista(request):
    dias = DiaEspecial.objects.order_by("-fecha")
    return render(
        request,
        "registro/dias_lista.html",
        {"dias": dias, "anio_actual": timezone.localdate().year},
    )


@staff_member_required
def dia_form(request, pk=None):
    dia = get_object_or_404(DiaEspecial, pk=pk) if pk else None

    # EDICIÓN: un solo día (formulario normal).
    if dia is not None:
        if request.method == "POST":
            fecha_anterior = dia.fecha
            form = DiaEspecialForm(request.POST, instance=dia)
            if form.is_valid():
                dia = form.save()
                _recalcular_tardanzas(fecha_anterior)
                if dia.fecha != fecha_anterior:
                    _recalcular_tardanzas(dia.fecha)
                messages.success(request, "Día especial actualizado.")
                return redirect("dias_lista")
        else:
            form = DiaEspecialForm(instance=dia)
        return render(request, "registro/dia_form.html", {"form": form, "dia": dia})

    # NUEVO: permite un rango de fechas (feriados de varios días).
    if request.method == "POST":
        form = RangoDiaEspecialForm(request.POST)
        if form.is_valid():
            from datetime import timedelta

            desde = form.cleaned_data["fecha_desde"]
            hasta = form.cleaned_data["fecha_hasta"] or desde
            tipo = form.cleaned_data["tipo"]
            hora = form.cleaned_data["hora_entrada_especial"]
            desc = form.cleaned_data["descripcion"]
            creados, existentes = 0, 0
            d = desde
            while d <= hasta:
                _, creado = DiaEspecial.objects.get_or_create(
                    fecha=d,
                    defaults={
                        "tipo": tipo,
                        "hora_entrada_especial": hora,
                        "descripcion": desc,
                    },
                )
                creados += 1 if creado else 0
                existentes += 0 if creado else 1
                _recalcular_tardanzas(d)
                d += timedelta(days=1)
            if creados:
                msg = f"Se registraron {creados} día(s) especial(es)."
                if existentes:
                    msg += f" {existentes} ya existían y se omitieron."
                messages.success(request, msg)
            else:
                messages.info(
                    request, "Todos los días del rango ya estaban registrados."
                )
            return redirect("dias_lista")
    else:
        form = RangoDiaEspecialForm()
    return render(request, "registro/dia_form.html", {"form": form, "dia": None})


def _pascua(anio):
    """Domingo de Pascua (algoritmo de Butcher) para calcular feriados móviles."""
    from datetime import date

    a = anio % 19
    b = anio // 100
    c = anio % 100
    d = b // 4
    e = b % 4
    ff = (b + 8) // 25
    g = (b - ff + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    ll = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * ll) // 451
    mes = (h + ll - 7 * m + 114) // 31
    dia = ((h + ll - 7 * m + 114) % 31) + 1
    return date(anio, mes, dia)


def _feriados_bolivia(anio):
    """Lista de (fecha, nombre) de los feriados de Bolivia para el año dado."""
    from datetime import date, timedelta

    pascua = _pascua(anio)
    feriados = [
        (date(anio, 1, 1), "Año Nuevo"),
        (date(anio, 1, 22), "Día del Estado Plurinacional de Bolivia"),
        (date(anio, 3, 6), "Aniversario de El Alto"),
        (pascua - timedelta(days=48), "Lunes de Carnaval"),
        (pascua - timedelta(days=47), "Martes de Carnaval"),
        (pascua - timedelta(days=2), "Viernes Santo"),
        (date(anio, 5, 1), "Día del Trabajo"),
        (pascua + timedelta(days=60), "Corpus Christi"),
        (date(anio, 6, 21), "Año Nuevo Andino Amazónico"),
        (date(anio, 7, 16), "Aniversario de La Paz"),
        (date(anio, 8, 6), "Día de la Independencia de Bolivia"),
        (date(anio, 11, 2), "Todos los Santos"),
        (date(anio, 12, 25), "Navidad"),
    ]
    feriados.sort(key=lambda x: x[0])
    return feriados


@staff_member_required
def dias_cargar_feriados(request):
    if request.method != "POST":
        return redirect("dias_lista")
    try:
        anio = int(request.POST.get("anio"))
    except (TypeError, ValueError):
        anio = timezone.localdate().year

    creados, existentes = 0, 0
    for fecha, nombre in _feriados_bolivia(anio):
        _, creado = DiaEspecial.objects.get_or_create(
            fecha=fecha,
            defaults={"tipo": DiaEspecial.FERIADO, "descripcion": nombre},
        )
        creados += 1 if creado else 0
        existentes += 0 if creado else 1
        _recalcular_tardanzas(fecha)

    msg = f"Feriados de Bolivia {anio}: se agregaron {creados}."
    if existentes:
        msg += f" {existentes} ya estaban registrados."
    messages.success(request, msg)
    return redirect("dias_lista")


@staff_member_required
def dias_recalcular_tardanzas(request):
    """Recalcula la tardanza de TODAS las asistencias (arregla datos viejos
    que no se actualizaron al agregar feriados o cambiar horarios)."""
    if request.method != "POST":
        return redirect("dias_lista")
    cambiadas = 0
    for a in Asistencia.objects.select_related("pasante"):
        nueva = (
            _calcular_tardanza(a.pasante, a.fecha, a.hora_entrada)
            if a.hora_entrada
            else 0
        )
        if nueva != a.tardanza_min:
            a.tardanza_min = nueva
            a.save(update_fields=["tardanza_min"])
            cambiadas += 1
    messages.success(
        request, f"Tardanzas recalculadas. Se corrigieron {cambiadas} registro(s)."
    )
    return redirect("dias_lista")


@staff_member_required
def dia_eliminar(request, pk):
    dia = get_object_or_404(DiaEspecial, pk=pk)
    if request.method == "POST":
        fecha = dia.fecha
        dia.delete()
        _recalcular_tardanzas(fecha)
        messages.success(request, "Día especial eliminado.")
        return redirect("dias_lista")
    return render(request, "registro/dia_eliminar.html", {"dia": dia})


# --------------------------- Reportes ---------------------------


def _asistencias_filtradas(request):
    hoy = timezone.localdate()
    desde = request.GET.get("desde") or hoy.replace(day=1).isoformat()
    hasta = request.GET.get("hasta") or hoy.isoformat()
    pasante_id = request.GET.get("pasante") or ""
    asistencias = (
        Asistencia.objects.select_related("pasante")
        .filter(fecha__gte=desde, fecha__lte=hasta)
        .order_by("-fecha", "pasante__nombre")
    )
    if pasante_id:
        asistencias = asistencias.filter(pasante_id=pasante_id)
    return asistencias, desde, hasta, pasante_id


@staff_member_required
def reportes(request):
    asistencias, desde, hasta, pasante_id = _asistencias_filtradas(request)

    if request.GET.get("export") == "excel":
        return _exportar_excel(asistencias, desde, hasta)

    total = asistencias.count()
    con_tardanza = asistencias.filter(tardanza_min__gt=0).count()
    minutos_tardanza = sum(a.tardanza_min for a in asistencias)

    return render(
        request,
        "registro/reportes.html",
        {
            "asistencias": asistencias,
            "pasantes": Pasante.objects.order_by("nombre"),
            "desde": desde,
            "hasta": hasta,
            "pasante_id": pasante_id,
            "total": total,
            "con_tardanza": con_tardanza,
            "minutos_tardanza": minutos_tardanza,
        },
    )


def _coord(lat, lng):
    if lat is None or lng is None:
        return ""
    return f"{lat:.5f}, {lng:.5f}"


def _exportar_excel(asistencias, desde, hasta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    ws.append(
        [
            "Pasante",
            "Identificador",
            "Área",
            "Fecha",
            "Entrada",
            "Salida",
            "Tardanza (min)",
            "Estado",
            "Ubicación ingreso",
            "Ubicación salida",
        ]
    )
    fuente = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="17803D")
    for celda in ws[1]:
        celda.font = fuente
        celda.fill = relleno
    for a in asistencias:
        ws.append(
            [
                a.pasante.nombre,
                a.pasante.identificador,
                a.pasante.area,
                a.fecha.strftime("%d/%m/%Y"),
                a.hora_entrada.strftime("%H:%M") if a.hora_entrada else "",
                a.hora_salida.strftime("%H:%M") if a.hora_salida else "",
                a.tardanza_min,
                a.get_estado_display(),
                _coord(a.lat_entrada, a.lng_entrada),
                _coord(a.lat_salida, a.lng_salida),
            ]
        )
    for i, ancho in enumerate([22, 14, 16, 12, 9, 9, 12, 18, 22, 22], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="asistencias_{desde}_a_{hasta}.xlsx"'
    )
    wb.save(response)
    return response


def admins_lista(request):
    admins = User.objects.filter(is_staff=True).order_by("username")
    return render(request, "registro/admins_lista.html", {"admins": admins})


@staff_member_required
def admin_form(request, pk=None):
    admin_user = get_object_or_404(User, pk=pk, is_staff=True) if pk else None

    if request.method == "POST":
        form = AdminUserForm(request.POST, instance=admin_user)
        if form.is_valid():
            password = form.cleaned_data["password"]

            if admin_user is None:
                if not password:
                    form.add_error(
                        "password",
                        "La contraseña es obligatoria al crear un administrador.",
                    )
                else:
                    obj = form.save(commit=False)
                    obj.is_staff = True
                    obj.set_password(password)
                    obj.save()
                    messages.success(request, f"Administrador «{obj.username}» creado.")
                    return redirect("admins_lista")
            else:
                # Protección: nadie se desactiva ni se quita superusuario a sí mismo.
                if admin_user.pk == request.user.pk:
                    form.instance.is_active = True
                    if (
                        not form.cleaned_data.get("is_superuser")
                        and admin_user.is_superuser
                    ):
                        messages.error(
                            request,
                            "No puedes quitarte tus propios permisos de superusuario.",
                        )
                        return render(
                            request,
                            "registro/admin_form.html",
                            {"form": form, "admin_user": admin_user},
                        )

                obj = form.save(commit=False)
                obj.is_staff = True
                if password:
                    obj.set_password(password)
                obj.save()
                messages.success(
                    request, f"Administrador «{obj.username}» actualizado."
                )
                return redirect("admins_lista")
    else:
        form = AdminUserForm(instance=admin_user)

    return render(
        request, "registro/admin_form.html", {"form": form, "admin_user": admin_user}
    )


@staff_member_required
def admin_eliminar(request, pk):
    """Da de baja o reactiva un administrador. Ya NO elimina la cuenta."""
    admin_user = get_object_or_404(User, pk=pk, is_staff=True)

    if admin_user.pk == request.user.pk:
        messages.error(request, "No puedes cambiar el estado de tu propia cuenta.")
        return redirect("admins_lista")

    if request.method == "POST":
        admin_user.is_active = not admin_user.is_active
        admin_user.save(update_fields=["is_active"])
        if admin_user.is_active:
            messages.success(
                request, f"Administrador «{admin_user.username}» reactivado."
            )
        else:
            messages.info(
                request,
                f"Administrador «{admin_user.username}» dado de baja. Ya no podrá iniciar sesión.",
            )
        return redirect("admins_lista")

    return render(request, "registro/admin_eliminar.html", {"admin_user": admin_user})


@staff_member_required
def pasante_reset_dispositivo(request, pk):
    pasante = get_object_or_404(Pasante, pk=pk)
    pasante.dispositivo_id = ""
    pasante.save(update_fields=["dispositivo_id"])
    messages.success(
        request,
        f"Dispositivo de «{pasante.nombre}» reiniciado. Podrá vincular un celular nuevo en su próxima marca.",
    )
    return redirect("pasantes_lista")


@staff_member_required
def planilla_word(request):
    """Planilla oficial en PDF (formato El Alto): encabezado/pie con imagen,
    sin columnas de FIRMA, recuadro de sello 2.7 x 3 cm, coordenadas en observaciones.
    (El nombre de la función queda igual para no cambiar urls.py.)"""
    import os
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    pasante_id = request.GET.get("pasante")
    if not pasante_id:
        messages.error(request, "Elige un pasante para generar su planilla.")
        return redirect("reportes")

    pasante = get_object_or_404(Pasante, pk=pasante_id)
    hoy = timezone.localdate()
    desde = request.GET.get("desde") or hoy.replace(day=1).isoformat()
    hasta = request.GET.get("hasta") or hoy.isoformat()

    asistencias = Asistencia.objects.filter(
        pasante=pasante, fecha__gte=desde, fecha__lte=hasta
    ).order_by("fecha")

    # Días especiales del rango, para anotarlos en Observaciones.
    dias_esp = {
        de.fecha: de
        for de in DiaEspecial.objects.filter(fecha__gte=desde, fecha__lte=hasta)
    }
    DIAS = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "DOMINGO"]

    base_img = os.path.join(os.path.dirname(__file__), "static", "img")
    img_h = os.path.join(base_img, "encabezado_el_alto.png")
    img_f = os.path.join(base_img, "pie_pagina.png")

    def marco(canvas, doc):
        W, H = LETTER
        hw = W - 24 * mm
        hh = hw * 0.1845
        if os.path.exists(img_h):
            canvas.drawImage(
                img_h, 12 * mm, H - 8 * mm - hh, width=hw, height=hh, mask="auto"
            )
        fw = W - 24 * mm
        fh = fw * 0.0753
        if os.path.exists(img_f):
            canvas.drawImage(img_f, 12 * mm, 5 * mm, width=fw, height=fh, mask="auto")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        topMargin=48 * mm,
        bottomMargin=24 * mm,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
    )

    titulo = ParagraphStyle(
        "t", fontName="Helvetica-Bold", fontSize=12, alignment=TA_CENTER, spaceAfter=10
    )
    info = ParagraphStyle(
        "i", fontName="Helvetica", fontSize=9, spaceAfter=3, leading=12
    )
    info_b = ParagraphStyle("ib", fontName="Helvetica-Bold", fontSize=8, spaceAfter=3)
    sello = ParagraphStyle(
        "s",
        fontName="Helvetica",
        fontSize=6.5,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#777777"),
    )
    hcab = ParagraphStyle(
        "h", fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER, leading=10
    )
    obs_st = ParagraphStyle(
        "o", fontName="Helvetica", fontSize=7, alignment=TA_LEFT, leading=8.5
    )

    d = timezone.datetime.fromisoformat(desde).strftime("%d/%m/%Y")
    h = timezone.datetime.fromisoformat(hasta).strftime("%d/%m/%Y")

    elems = []
    elems.append(
        Paragraph("<u>PLANILLA DE ASISTENCIA TRABAJO DIRIGIDO O PASANTÍA</u>", titulo)
    )

    # Estilos para la firma del supervisor (líneas + etiqueta, centrados)
    firma_lineas = ParagraphStyle(
        "fl", fontName="Helvetica", fontSize=10, alignment=TA_CENTER, spaceAfter=1
    )
    firma_lbl = ParagraphStyle(
        "flb", fontName="Helvetica-Bold", fontSize=8, alignment=TA_CENTER, spaceAfter=3
    )

    # Recuadro del sello: cuadro vacío (2.7 x 3 cm) con la etiqueta DEBAJO.
    caja_sello = Table([[""]], colWidths=[2.7 * cm], rowHeights=[3 * cm])
    caja_sello.setStyle(
        TableStyle([("BOX", (0, 0), (0, 0), 0.8, colors.HexColor("#888888"))])
    )
    sello_cell = [
        caja_sello,
        Spacer(1, 3),
        Paragraph("SELLO DE LA DEPENDENCIA", sello),
    ]

    info_cell = [
        Paragraph(f"<b>DEPENDENCIA:</b> {pasante.area}", info),
        Spacer(1, 16),
        Paragraph(
            "_______________ &nbsp;&nbsp;&nbsp;&nbsp; _______________", firma_lineas
        ),
        Paragraph("<b>FIRMA Y SELLO DEL SUPERVISOR</b>", firma_lbl),
        Spacer(1, 12),
        Paragraph(
            f"<b>NOMBRE DEL PASANTE:</b> {pasante.nombre} &nbsp;&nbsp; <b>C.I.:</b> {pasante.ci}",
            info,
        ),
        Paragraph(f"<b>FECHA DESDE EL MES DE:</b> {d} &nbsp; <b>AL:</b> {h}", info),
    ]
    htbl = Table([[info_cell, sello_cell]], colWidths=[132 * mm, 54 * mm])
    htbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ]
        )
    )
    elems.append(htbl)
    elems.append(Spacer(1, 10))

    # Tabla SIN columnas de FIRMA (7 columnas)
    cab = [
        Paragraph(t, hcab)
        for t in [
            "Nº",
            "DIA",
            "FECHA",
            "HORA DE<br/>INGRESO",
            "HORA DE<br/>SALIDA",
            "OBSERVACIONES",
            "VºBº<br/>SUPERVISOR",
        ]
    ]
    data = [cab]

    for i, a in enumerate(asistencias, start=1):
        obs = []
        if a.lat_entrada is not None:
            obs.append("Entrada marcada dentro de las instalaciones.")
        if a.lat_salida is not None:
            obs.append("Salida marcada dentro de las instalaciones.")
        if a.tardanza_min:
            obs.append(f"Retraso: {a.tardanza_min} min (superó la tolerancia).")
        esp = dias_esp.get(a.fecha)
        if esp:
            if esp.tipo == DiaEspecial.FERIADO:
                txt = "Feriado"
                if esp.descripcion:
                    txt += f": {esp.descripcion}"
                obs.append(txt + ".")
            else:
                txt = "Día especial"
                if esp.hora_entrada_especial:
                    txt += f" (entrada {esp.hora_entrada_especial.strftime('%H:%M')})"
                if esp.descripcion:
                    txt += f": {esp.descripcion}"
                obs.append(txt + ".")
        data.append(
            [
                str(i),
                DIAS[a.fecha.weekday()],
                a.fecha.strftime("%d/%m/%y"),
                a.hora_entrada.strftime("%H:%M") if a.hora_entrada else "",
                a.hora_salida.strftime("%H:%M") if a.hora_salida else "",
                Paragraph("<br/>".join(obs), obs_st),
                "",
            ]
        )
    for _ in range(max(0, 16 - len(asistencias))):
        data.append([""] * 7)

    tbl = Table(
        data,
        colWidths=[9 * mm, 24 * mm, 20 * mm, 24 * mm, 24 * mm, 58 * mm, 27 * mm],
        repeatRows=1,
    )
    tbl.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, -1), 8.5),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elems.append(tbl)

    doc.build(elems, onFirstPage=marco, onLaterPages=marco)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="planilla_{pasante.identificador}.pdf"'
    )
    return response
